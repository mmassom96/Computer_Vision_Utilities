import shutil
import os
import openpyxl
from openpyxl import load_workbook

wb = load_workbook("/media/matthew/Matt's Backup HDD/Beating_Roulette_Data/spins.xlsx")

sh = wb.active

BINARY = True

SOURCE = "/media/matthew/Matt's Backup HDD/Beating_Roulette_Data/Dataset_master/single_spins/"
DESTINATION = "/media/matthew/Matt's Backup HDD/Beating_Roulette_Data/Dataset_master/spins_named/"

row = 2
column = 1

if  BINARY is False:
    DESTINATION = DESTINATION + "true_outcome/"
    try:
        if not os.path.exists(DESTINATION):
            os.makedirs(DESTINATION)
    except OSError:
        print ('Error: Creating directory: ', DESTINATION)
    
    while(sh.cell(row, 1).value != None):
        # videoName_resolution_frameRate_xCenterPoint_yCenterPoint_outcome.MP4
        formatted_name = str(sh.cell(row, 1).value).replace(".MP4", "_") + str(sh.cell(row, 2).value) + "_" + str(sh.cell(row, 3).value) + "fps_" + str(sh.cell(row, 4).value) + "_" + str(sh.cell(row, 5).value) + "_" + str(sh.cell(row, 6).value) + ".MP4"
        src = SOURCE + str(sh.cell(row, 1).value)
        dst = DESTINATION + formatted_name
        shutil.copy(src ,dst)
        row += 1

else:
    DESTINATION = DESTINATION + "binary_outcome/"
    try:
        if not os.path.exists(DESTINATION):
            os.makedirs(DESTINATION)
    except OSError:
        print ('Error: Creating directory: ', DESTINATION)
    
    zero = ["1", "13", "36", "24", "3", "15", "34", "22", "5", "17", "32", "20", "7", "11", "30", "26", "9", "28", "0"]
    one = ["2", "14", "35", "23", "4", "16", "33", "21", "6", "18", "31", "19", "8", "12", "29", "25", "10", "27", "00"]
    
    while(sh.cell(row, 1).value != None):
        # videoName_resolution_frameRate_xCenterPoint_yCenterPoint_binaryOutcome.MP4
        if str(sh.cell(row, 6).value) in zero:
            binaryOutcome = "0"
        elif str(sh.cell(row, 6).value) in one:
            binaryOutcome = "1"
        else:
            binaryOutcome = "NULL"
        formatted_name = str(sh.cell(row, 1).value).replace(".MP4", "_") + str(sh.cell(row, 2).value) + "_" + str(sh.cell(row, 3).value) + "fps_" + str(sh.cell(row, 4).value) + "_" + str(sh.cell(row, 5).value) + "_" + binaryOutcome + ".MP4"
        src = SOURCE + str(sh.cell(row, 1).value)
        dst = DESTINATION + formatted_name
        shutil.copy(src ,dst)
        row += 1